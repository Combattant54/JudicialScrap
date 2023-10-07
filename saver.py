from openpyxl import Workbook
from openpyxl.cell import Cell
import os
import logging
logger = logging.getLogger(__name__)

class Saver():
    ALPHABET = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "w", "X", "Y", "Z",
    ]
    
    def __init__(self, name, folders, columns_map) -> None:
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = str(name)
        self.columns = dict(columns_map)
        
        self.folders = folders

        self.line = 1
        for k, v in self.columns.items():
            target = f"{self._LETTERS[k] if isinstance(k, int) else k}{self.line}"
            self.ws[target] = v
        self.increase_line()
            
        self.comp_name = ""
        self.comp_date = ""
        self.count_name = ""
        self.cit_name = ""
        self.add_inf = ""
        self.current_category = ""
        
        
        result = self.ws["A1:Z1"]
        for e in result[0]:
            assert isinstance(e, Cell)
            #print(e.coordinate, e.value)
    
    def save_categorie(self, name):
        self.current_category = name
    
    def end_current_categorie(self, number):
        self.save_personne(additional_information=str(number))
        self.current_category = ""

    def increase_line(self):
        self.line += 1
    
    def end_competition_saving(self):
        self.ws[f"A{self.line}"] = self.comp_name
        self.ws[f"B{self.line}"] = self.comp_date
        self.ws[f"C{self.line}"] = self.count_name
        self.ws[f"D{self.line}"] = self.cit_name
        self.ws[f"L{self.line}"] = ""
        self.comp_name = ""
        self.comp_date = ""
        self.count_name = ""
        self.cit_name = ""
        self.add_inf = list()

        self.end_current_categorie("")
        self.increase_line()

    def save_competition(self, competition_name, competition_dates, country_name, city_name="", additional_information=""):
        if self.comp_name:
            self.end_competition_saving()
        self.comp_name = competition_name
        self.comp_date = competition_dates
        self.count_name = country_name
        self.cit_name = city_name
        self.add_inf = list(additional_information)

    def add_informations(self, informations):
        self.add_inf.append(informations)
    
    def save_personne(self, name="", nationality="", rank="", rank_description="", category="", other_prices="", competition_category=None, additional_information=""):
        try:
            self.ws[f"A{self.line}"] = self.comp_name
            self.ws[f"B{self.line}"] = self.comp_date
            self.ws[f"C{self.line}"] = self.count_name
            self.ws[f"D{self.line}"] = self.cit_name
            self.ws[f"E{self.line}"] = self._CATEGORIES[int(category)] if str(category).isnumeric() else ""
            self.ws[f"F{self.line}"] = name, False
            self.ws[f"G{self.line}"] = rank
            self.ws[f"H{self.line}"] = rank_description
            if len(str(nationality)) <= 2:
                nationality = ""
            self.ws[f"I{self.line}"] = nationality
            if self.current_category != "":
                if competition_category is None:
                    competition_category = self.current_category
            self.ws[f"J{self.line}"] = competition_category
            self.ws[f"K{self.line}"] = other_prices
            self.ws[f"L{self.line}"] = additional_information
            self.increase_line()
        except Exception as e:
            logger.exception(e)
        
    
    def save(self, file_name):
        file_name = file_name if file_name.endswith(".xlxs") else file_name+".xlxs"
        file_path = os.path.join(*self.folders, file_name)
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            print(e)
        finally:
            self.wb.save(file_path)
